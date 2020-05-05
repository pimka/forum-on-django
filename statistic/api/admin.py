import datetime

import openpyxl
from django.contrib import admin
from django.http import HttpResponse

from .models import StatisicModel


def export_data(modeladmin, request, queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"report-{datetime.datetime.now().strftime('%Y-%m-%d')}"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ['user_uuid', 'datetime', 'operation', 'before', 'after']

    for col, header in enumerate(headers):
        ws.cell(row=1, column=col+1, value=header)

    for row_n, row in enumerate(queryset):
        fields = (row.user_uuid, row.datetime_operation, row.operation, row.before_changes, row.after_changes)
        for col, value in enumerate(fields):
            ws.cell(row=row_n+2, column=col+1, value=value)
    
    wb.save(response)
    return response

export_data.short_description = 'Export to file'

class StatisticAdmin(admin.ModelAdmin):
    list_display = ['user_uuid', 'operation', 'datetime_operation']
    ordering = ['-datetime_operation']
    actions = [export_data]

admin.site.register(StatisicModel, StatisticAdmin)
